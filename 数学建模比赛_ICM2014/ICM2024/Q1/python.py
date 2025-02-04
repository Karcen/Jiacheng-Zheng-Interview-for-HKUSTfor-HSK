import pandas as pd
import numpy as np
from scipy.optimize import minimize, stats
import os
import networkx as nx
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import sys
# Question 1
def process_excel_to_csv(workbook_path):
    # Function to convert Excel sheets to CSV file # ... 
def merge_csv_files(generated_csv_paths):
    # Function to merge CSV files # ... 
# Example usage
workbook_path = 'Problem D Great Lakes.xlsx'
generated_csv_paths = process_excel_to_csv(workbook_path)
merged_df = merge_csv_files(generated_csv_paths)
merged_df[['Source', 'Year', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']].to_csv('data.csv', index=None)
merged_df_cleaned = pd.read_csv('data.csv', na_values='---')
# Example usage (continued)
ideal_levels = {
    'Superior': 183.35,
    'Michigan-Huron': 176.33,
    'St. Clair': 175.10,
    'Erie': 174.28,
    'Ontario': 74.83
}
optimized_data = optimize_lake_levels(merged_df_cleaned)
pd.DataFrame(optimized_data, columns=['Month', 'Erie', 'Michigan', 'Ontario', 'Clair', 'Superior']).to_csv('Q1.csv', index=False)
# Question 2
G = nx.DiGraph()
nodes = ['Superior', 'Michigan&Huron', 'St. Clair', 'Erie', 'Ontario']
G.add_nodes_from(nodes)
edges_with_rivers = [
    ('Superior', 'Michigan&Huron', {'river': 'St. Marys River'}),
    ('Michigan&Huron', 'St. Clair', {'river': 'St. Clair River'}),
    ('St. Clair', 'Erie', {'river': 'Detroit River'}),
    ('Erie', 'Ontario', {'river': 'Niagara River'})
]
G.add_edges_from(edges_with_rivers)
pos = nx.circular_layout(G)
plt.figure(figsize=(12, 10))
node_labels = {node: node for node in nodes}
nx.draw_networkx_labels(G, pos, labels=node_labels, font_size=15, font_color='black', font_weight='bold')
edge_labels = {(u, v): data['river'] for u, v, data in G.edges(data=True)}
nx.draw_networkx_edge_labels(G, pos, edge_labels=edge_labels, font_size=12, font_color='red')
nx.draw_networkx_edges(G, pos, edge_color='darkblue', width=3, arrowstyle='->', arrowsize=40, style='dashed')
nx.draw_networkx_nodes(G, pos, node_size=5000, node_color='skyblue')
plt.legend(['Water Flow'], loc='upper right', fontsize=12)
plt.axis("off")
plt.show()
data = pd.read_csv('data.csv')
data_info = data.info()
data_head = data.head()
unique_sources = data['Source'].unique()
lake_means = data.groupby('Source').mean().drop(columns='Year')
def calculate_total_difference(ideal_levels, actual_levels):
    return abs(actual_levels - ideal_levels).sum().sum()
ideal_water_levels = lake_means.mean(axis=1)
print("Ideal Water Levels:")
print(ideal_water_levels)
lake_max = data.groupby('Source').max().drop(columns='Year')
def calculate_total_difference(ideal_levels, actual_levels):
    return abs(actual_levels - ideal_levels).sum().sum()
ideal_water_max = lake_max.max(axis=1)
print("Ideal Maximum Water Levels:")
print(ideal_water_max)
def calculate_monthly_flows(data, month):
    rivers = ['St. Mary\'s River - Flow ', 'St. Clair River - Flow ', 'Detroit River - Flow ', 'Niagara River - Flow at Buffalo']
    return [data[data['Source'] == river][month].mean(numeric_only=True) for river in rivers]
def adjust_initial_flows(monthly_flows):
    return np.array(monthly_flows)
def objective_function(flows, ideal_levels, max_capacity):
    return sum(abs(ideal_levels[node] - flows[i] / max_capacity[river]) for i, (node, river) in enumerate(zip(ideal_levels.keys(), max_capacity.keys())))
def optimize_water_levels(data, ideal_levels, max_capacity, months):
    result_dataframe = []
    for month in months:
        monthly_flows = calculate_monthly_flows(data, month)
        initial_flows = adjust_initial_flows(monthly_flows)
        bounds = [(0, max_capacity[river]) for river in max_capacity]
        optimization_result = minimize(objective_function, initial_flows, args=(ideal_levels, max_capacity), bounds=bounds, method='SLSQP')
        optimized_levels = optimization_result.x
        result_dataframe.append([month] + list(optimized_levels))
    return result_dataframe
months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
result_dataframe = optimize_water_levels(data, ideal_water_levels, max_flow_capacity, months)
for row in result_dataframe:
    print(row)
pd.DataFrame(newdf, columns=['month', 'Marys', 'Clair', 'Detroit', 'Niagara'])
pd.DataFrame(newdf, columns=['month', 'Marys', 'Clair', 'Detroit', 'Niagara']).to_csv('Q2.csv')

# Question 3
# Water Flow Network Visualization
data = pd.read_csv('data.csv')
def visualize_water_flow_network():
    # NetworkX DiGraph creation
    water_flow_network = nx.DiGraph()
    lake_nodes = ['Superior', 'Michigan&Huron', 'St. Clair', 'Erie', 'Ontario', 'Soo Locks', 'Moses-Saunders Dam']
    water_flow_network.add_nodes_from(lake_nodes)
    edges_with_rivers = [
        # ... (edges with rivers and their capacities)
    ]
    water_flow_network.add_edges_from(edges_with_rivers)
    node_positions = nx.spring_layout(water_flow_network)
visualize_water_flow_network()
# Optimization Process
optimized_results = []
ideal_levels = {'Superior': 183.35, 'Michigan&Huron': 176.33, 'St. Clair': 175.10, 'Erie': 174.28, 'Ontario': 74.83}
max_flow_capacity = {'St. Marys River': 3191.308305, 'St. Clair River': 7011.250543, 'Detroit River': 7676.696374, 'Niagara River': 8070.0}
for i in range(1, 11):
    dam_impact_soo_locks += i / 10
    for month in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']:
        # ... (calculations and optimization process)
        optimized_results.append([i, month, result.fun])
pd.DataFrame(optimized_results, columns=['dam_impact_SooLocks', 'month', 'Deviation']).to_csv('Q3_1.csv', index=None)

# Question 4
# Processing Excel Sheet
def process_excel_sheet(sheet):
    # ... (extracting filename, creating CSV, and storing paths)
def read_and_clean_data():
    # ... (reading and merging data from generated CSVs)
def replace_and_fill_missing(merged_df_cleaned):
    # ... (filling missing values with means)
def create_water_flow_network():
    # ... (NetworkX DiGraph creation with lakes, dams, and rivers)
def optimize_water_levels():
    # ... (optimization process for water levels)
def perform_anova():
    # ... (performing ANOVA on different scenarios)
# Optimization Process
workbook_path = 'Problem_D_Great_Lakes.xlsx'
workbook = load_workbook(workbook_path)
generated_csv_paths = []
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    process_excel_sheet(sheet)
merged_data = read_and_clean_data()
merged_data.to_csv('data.csv', index=None)
merged_data_cleaned = pd.read_csv('data.csv', na_values='---')
merged_data_cleaned = replace_and_fill_missing(merged_data_cleaned)
create_water_flow_network()
optimize_results = optimize_water_levels()
optimize_results.to_csv('optimized_results.csv', index=None)
perform_anova()